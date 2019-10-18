# coding=utf-8  
"""
@author: mirrorChen
@license: (C) Copyright 2011-2018, mirror personal Limited.
@contact: chenjingxu3@dafycredit.com
@software: JY_Android_AT
@file: main.py
@time: 2019/6/26 17:05
@desc: 
"""

import xlsxwriter, xlwt
import js2py,execjs
from appium import webdriver
import unittest, time
import random
from mysql_pubtwo import MysqlUtiltwo
from oracle_pub import OracleUtil
from excel_pub import ExcelUtil
from InterfaceClass import APITest
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import uiautomator2 as u2

from API_Para import API_Parameters
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
class Method():
    def __init__(self, driver):
        self.driver = driver

    def findItem(self,el):
        time.sleep(1)
        source = self.driver.page_source
        if el in source:
            return True
        else:
            return False


    def createName(self):
        a1 = ['张', '金', '李', '王', '赵', '陈', '胡', '莫', '卫', '韦', '霍', '贾', '周', '陆']
        a2 = ['玉', '明', '龙', '芳', '军', '玲', '镜', '元', '都', '红', '绑', '问', '有', '那', '及', '公', '的', '龙', '格', '达', '离','家', '个','极','开']
        a3 = ['来', '立', '玲', '', '国', '及', '订', '额', '发', '哥', '试', '还', '吧', '饿', '运', '扣', '流', '考', '岭', '嗯', '玩','解', '密', '点', '来', '放', '哦']
        name = "测试" + random.choice(a1) + random.choice(a2) + random.choice(a3)
        return name

    def insertCell(self, oPath, TsheetName, is_origin, row, col, contens, sPath=None):
        #插入列数据。
        wb = load_workbook(oPath)
        if is_origin == True:
            wsO = wb[TsheetName]
            wsO.cell(int(row), int(col)).value = contens
            wb.save(oPath)
            print("保存成功！")
        elif is_origin == False:
            wbT = load_workbook(sPath)
            wsO = wbT[TsheetName]
            wsO.cell(int(row), int(col)).value = contens
            wbT.save(sPath)
            print("保存成功！")
        else:
            print("请输入‘True’或者 ‘False’" )
            return


class createData(unittest.TestCase):
    def setUp(self):
        #初始化信息
        self.Excel = ExcelUtil("./userinfo.xlsx", "ai")
        self.MDB = MysqlUtiltwo()
        self.ODB = OracleUtil()
        self.API = APITest()
        self.API_Para = API_Parameters()

        #定义机器信息  小米6, nova2，魅族note2 vivo
        # desired_caps = {}
        # desired_caps['platformName'] = 'Android'
        # desired_caps['platformVersion'] = '7.1.2'  #'8.0.0' 7.1.2
        # desired_caps['deviceName'] = '721QEBR925Q45'  #'RJCDU17609004486'  fc323ca3  c9edca71 721QEBR925Q45 bb76fd37
        # desired_caps['appPackage'] = 'com.giveu.shoppingmall'
        # desired_caps['appActivity'] = 'com.giveu.shoppingmall.index.view.activity.SplashActivity'
        # desired_caps['unicodeKeyboard'] = 'true'
        # desired_caps['resetKeyboard'] = 'true'
        # desired_caps['automationName'] = 'uiautomator2'
        # self.driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', desired_caps)  #10.10.72.188
        self.driver = u2.connect("721QEBR925Q45")
        self.Method = Method(self.driver)
        self.driver.wait_timeout(30.0)


    def test_create_data(self):
        # 造数据
        info = self.Excel.next()
        i = 0
        j = len(info) - 1
        registered = [] #记录已注册用户
        while j >= 0:
            if info[j]["REGISTER"] == "注册成功":
                registered.append(info.pop(j)) #已注册的则剔除,并记录到一个列表中
            j = j - 1

        width = self.driver.get_window_size()['width']  # 获取屏幕的宽度
        height = self.driver.get_window_size()['height']  # 获取屏幕的高度
        while i < len(info):
            try:
                self.ID = info[i]["ID"]
                self.PHONE = info[i]["PHONE"]
                self.ACCOUNT_NO = info[i]["ACCOUNT_NO"]
                self.IDENT = info[i]["IDENT"]
                self.REGISTER = info[i]["REGISTER"]
                self.REMARK = info[i]["REMARK"]

                uName = self.Method.createName()
                print("提单人姓名：", uName)
                print("提单手机号：", self.PHONE)
                print("提单身份证：", self.IDENT)
                print("提单银行卡：", self.ACCOUNT_NO)

                # 造数据：到信用授权通过.
                self.driver(textContained="允许").click()
                self.driver(textContained="允许").click()
                # if self.Method.findItem("允许"):
                #     self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"允许\")").click()
                #     self.driver.find_elements_by_android_uiautomator("new UiSelector().text(\"允许\")")[0].click()
                #     # self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"始终允许\")").click()
                #     print("授权成功")
                # 注册
                # 注册 --跳转注册页面
                print("开始注册")
                self.driver.find_elements_by_id("bottom_icon")[4].click()  # 点击我的
                # self.driver.find_elements(By.ID,"com.giveu.shoppingmall:id/bottom_icon")[4].click()
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_login").click()
                time.sleep(0.5)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_register").click()
                time.sleep(0.5)

                # 注册--注册账号页面
                print("注册账号")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_content")[0].send_keys(self.PHONE)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_send_code").click()
                time.sleep(2)  # 等两秒再去填写验证码
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_content")[1].send_keys("123456")
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_next").click()
                time.sleep(0.5)
                # 注册--设置密码页面
                print("设置密码")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_content")[0].send_keys("abc12345")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_content")[1].send_keys("abc12345")
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_complete").click()
                print("注册成功")
                # EXCEL中标记注册成功
                self.Method.insertCell("./userinfo.xlsx", "ai", True, int(self.ID) + 1, 5, "注册成功")
                time.sleep(1)
                # 登录  --注册之后不需要填写账号
                print("登录账号")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_content")[1].send_keys("abc12345")
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_login").click()
                time.sleep(2)
                # self.driver.find_element_by_id("tv_firstBtn").click()
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"取消\")").click()
                print("登陆成功！")
                # self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"取消\")").click()
                time.sleep(0.5)

                # 取现部分----------------------------------------------------------------------------------------------------
                print("进入取现流程")

                # self.driver.wait_activity("com.giveu.shoppingmall.index.view.activity.MainActivity")
                # self.driver.find_elements_by_id("com.giveu.shoppingmall:id/bottom_layout")[1].click()  # 点击取现
                self.driver.tap([(((width / 5) + (width / 5)) - 50, height - 50), (((width / 5) + (width / 5)) - 40, height - 40)],500)
                time.sleep(1)
                #self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"立即申请\")").click()
                self.driver.find_element_by_id("tv_operate").click()
                # 调用接口，插入活体图片数据，跳过活体
                getparams = ''
                sql = "select access_token from t_user_token tu where user_name="+str(self.PHONE)+" order by create_time desc limit 1"
                self.tk = self.MDB.mysql_getstring(sql)
                sql = "select id from t_user where login_name=" + str(self.PHONE)
                userid = self.MDB.mysql_getstring(sql)
                postparams = self.API_Para.skiplivingbody(userid)
                user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'

                tokens = 'Authorization: Bearer ' + str(self.tk)
                # 添加头信息
                headers = {'User-Agent': user_agent,
                           'Content-Type': 'application/json',
                           'token': tokens
                           }
                retInfo = self.API.apicall("POST",
                                           "http://testdafyshop.dafycredit.cn/v1/cashloan/applyInfoRecord/substepSave",
                                           getparams,
                                           postparams,
                                           headers
                                           )
                print(retInfo)
                self.driver.find_element_by_id("tv_next").click()

                # 取现部分--授权
                if self.Method.findItem("允许"):
                    self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"允许\")").click()
                    self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"允许\")").click()
                    # self.driver.find_elements_by_android_uiautomator("new UiSelector().text(\"允许\")")[0].click()

                # 取现部分 --绑定身份证信息
                # 正面
                print("身份证正面")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/iv_pocket_money_add")[0].click()
                time.sleep(1)
                self.driver.find_element_by_id("com.meizu.media.camera:id/shutter_btn").click()  # 魅族note6
                #self.driver.find_element_by_id("v9_shutter_button_internal").click()  #小米6
                # self.driver.find_element_by_id("com.huawei.camera:id/shutter_button").click() #华为nova2
                time.sleep(3)
                self.driver.find_element_by_id("com.meizu.media.camera:id/btn_done").click()  # 魅族note6
                #self.driver.find_element_by_id("inten_done_apply").click() #小米6
                # self.driver.find_element_by_id("com.huawei.camera:id/done_button").click()#华为nova2
                # 反面
                time.sleep(2)
                print("身份证反面")
                self.driver.swipe(460, 950, 460, 200)
                time.sleep(0.5)
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/iv_pocket_money_add")[0].click()
                time.sleep(2)
                self.driver.find_element_by_id("com.meizu.media.camera:id/shutter_btn").click()  # 魅族note6
                # self.driver.find_element_by_id("com.huawei.camera:id/shutter_button").click()#华为nova2
                # time.sleep(6)
                # self.driver.find_element_by_id("com.huawei.camera:id/done_button").click()#华为nova2
                #self.driver.find_element_by_id("v9_shutter_button_internal").click()#小米6
                time.sleep(3)
                self.driver.find_element_by_id("com.meizu.media.camera:id/btn_done").click()  # 魅族note6
                # time.sleep(2)
                #self.driver.find_element_by_id("inten_done_apply").click()#小米6
                # 手持
                time.sleep(2)
                print("手持身份证")
                self.driver.swipe(460, 950, 460, 200)
                time.sleep(0.5)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/iv_pocket_money_add").click()
                time.sleep(2)
                self.driver.find_element_by_id("com.meizu.media.camera:id/shutter_btn").click()  # 魅族note6
                # self.driver.find_element_by_id("com.huawei.camera:id/shutter_button").click()华为nova2
                # time.sleep(6)
                # self.driver.find_element_by_id("com.huawei.camera:id/done_button").click()华为nova2
                #self.driver.find_element_by_id("v9_shutter_button_internal").click()#小米6
                time.sleep(2)
                self.driver.find_element_by_id("com.meizu.media.camera:id/btn_done").click()  # 魅族note6
                #self.driver.find_element_by_id("inten_done_apply").click()#小米6
                # 填写身份证信息
                time.sleep(2.5)
                print("身份证信息")
                self.driver.swipe(460, 950, 460, 150)
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[0].clear()  # 姓名
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[0].send_keys(uName)
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[1].clear()  # 身份证号码
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[1].send_keys(self.IDENT)
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[2].clear()  # 户籍地址
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[2].send_keys("福中一路地铁大厦17楼")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[3].clear()  # 发证机关
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[3].send_keys("福田区公安局")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/tv_text_right")[0].click()
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"京族\")").click()
                self.driver.find_elements_by_id("tv_text_right")[1].click()
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"福建\")").click()
                time.sleep(0.3)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"莆田市\")").click()
                time.sleep(0.3)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"仙游县\")").click()
                time.sleep(0.3)
                self.driver.find_elements_by_id("tv_text_right")[2].click()  # 有效日期起
                time.sleep(0.5)
                self.driver.swipe(180, 1384, 180, 1700)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"确定\")").click()
                time.sleep(0.3)
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/tv_text_right")[3].click()  # 有效日期止
                time.sleep(0.3)
                self.driver.swipe(180, 1384, 180, 1000)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"确定\")").click()
                time.sleep(0.5)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_next").click()
                time.sleep(2)

                # 基本信息
                # 个人
                print("个人信息")
                self.driver.find_elements_by_id("tv_text_right")[0].click()  # 居住地址
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"福建\")").click()
                time.sleep(0.3)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"莆田市\")").click()
                time.sleep(0.3)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"仙游县\")").click()
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[0].send_keys(
                    "福中一路地铁大厦17楼")  # 详细地址
                self.driver.find_elements_by_id("tv_text_right")[1].click()  # 学历
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"小学\")").click()
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[1].send_keys(
                    "123456@qq.com")  # 邮箱
                self.driver.find_elements_by_id("tv_text_right")[2].click()  # 婚姻状态
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"未婚\")").click()
                self.driver.find_elements_by_id("tv_text_right")[3].click()  # 月收入
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"3000-4000\")").click()
                # 公司
                print("公司信息")
                self.driver.find_elements_by_id("et_text_right")[2].send_keys("深圳即融科技有限公司")  # 公司名称
                self.driver.swipe(460, 1080, 460, 100)
                time.sleep(1.5)
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[0].send_keys(
                    "0755-86511248")  # 公司电话
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/tv_text_right")[0].click()
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"司机\")").click()
                # 家庭联系人
                print("家庭联系人信息")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[1].send_keys("熊大")  # 姓名
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[2].send_keys(
                    "13410342889")  # 手机号码
                self.driver.find_elements_by_id("tv_text_right")[1].click()  # 与本人关系
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"兄弟\")").click()
                # 其它联系人
                print("其它联系人信息")
                self.driver.find_elements_by_id("et_text_right")[3].send_keys("熊二")  # 姓名
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/et_text_right")[4].send_keys(
                    "13410342888")  # 手机号码
                self.driver.find_elements_by_id("tv_text_right")[2].click()  # 与本人关系
                time.sleep(1)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"同学\")").click()
                # 提交下一步
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_next").click()
                time.sleep(1)

                # 绑定银行卡
                print("绑定银行卡")
                # self.driver.activate_ime_engine("com.baidu.input_huawei/.ImeService")
                self.driver.find_elements_by_id("et_text_right")[1].send_keys(self.ACCOUNT_NO)  # 银行卡号
                # self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"请填写银行卡卡号\")").send_keys(self.ACCOUNT_NO)
                # self.driver.activate_ime_engine("io.appium.android.ime/.UnicodeIME") 请填写银行预留手机号
                self.driver.find_elements_by_id("et_text_right")[2].click()
                time.sleep(1)
                self.driver.find_elements_by_id("et_text_right")[2].send_keys(self.PHONE)  # 预留手机号
                time.sleep(0.5)
                self.driver.find_element_by_id("tv_send_code").click()
                time.sleep(2)
                self.driver.find_element_by_id("et_send_code").send_keys("123456")
                # 提交下一步
                self.driver.find_element_by_id("tv_next").click()
                time.sleep(0.3)
                print("绑定银行卡成功")
                # 信用提升
                print("信用提升")
                self.driver.find_elements_by_id("com.giveu.shoppingmall:id/bt_credit")[0].click()
                time.sleep(2)
                sql = "update dafy_sales.EXTERNAL_FS_TASK set task_id='fsd4354332jk3123d',is_get_report=1 where name='" + uName + "'"
                self.ODB.oracle_sql(sql)
                time.sleep(1)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/top_left_image").click()
                #self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_refresh").click()  # 刷新
                time.sleep(1.5)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_next").click()
                print("授权成功")
                # EXCEL中标记授权成功
                self.Method.insertCell("./userinfo.xlsx", "ai", True, int(self.ID) + 1, 6, "授权成功")
                # EXCEL中标记激活成功
                self.Method.insertCell("./userinfo.xlsx", "ai", True, int(self.ID) + 1, 6, "已激活")
                self.Method.insertCell("./userinfo.xlsx", "ai", True, int(self.ID) + 1, 9, uName)
                i = i + 1
                print("第 " + str(i) + " 个账号数据已创建成功")
                time.sleep(1)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/top_left_image").click()
                time.sleep(0.5)
                # self.driver.find_element_by_id("com.giveu.shoppingmall:id/top_left_image").click()
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"我知道了\")").click()
                time.sleep(0.5)
                # 关闭因设备限制而弹出的弹窗

                #提现申请
                #do sth

            except Exception as e:
                print(e)
            finally:
                # 退出
                self.driver.press_keycode(4)
                self.driver.press_keycode(4)
                self.driver.press_keycode(4)
                print("退出登录")
                self.driver.find_elements_by_id("bottom_icon")[4].click()  # 点击我的
                time.sleep(0.5)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/iv_avatar").click()
                time.sleep(0.5)
                self.driver.swipe(460, 1000, 460, 100)
                self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_finish").click()
                time.sleep(0.5)
                self.driver.find_element_by_android_uiautomator("new UiSelector().text(\"确定\")").click()
                time.sleep(1)
                print("退出成功")


    def tearDown(self):
        self.driver.quit()


if __name__ == "__main__":
    unittest.main()












