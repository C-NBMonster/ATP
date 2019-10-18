# coding:utf-8
import xlrd#,xlwt,xlsxwriter
# from xlutils.copy import copy
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

class ExcelUtil():

    def __init__(self, excelPath, sheetName):
        self.data = xlrd.open_workbook(excelPath) #formatting_info=True保留格式
        self.table = self.data.sheet_by_name(sheetName)
        # get titles
        self.row = self.table.row_values(0)
        self.col = self.table.col_values(0)
        # get rows number
        self.rowNum = self.table.nrows
        # get columns number
        self.colNum = self.table.ncols
        # the current column
        self.curRowNo = 1
        self .cPath = excelPath
        self.sName  = sheetName

    def hasNext(self):
        if self.rowNum == 0 or self.rowNum <= self.curRowNo :
            return False
        else:
            return True

    def next(self):
        r = []
        while self.hasNext():
            s = {}
            col = self.table.row_values(self.curRowNo)
            i = self.colNum
            for x in range(i):
                s[self.row[x]] = col[x]
            r.append(s)
            self.curRowNo += 1
        return r

    def rowlist(self, i):
        # 按行读取存为list,去除空字符
        rowlist = self.table.row_values(i)
        n = rowlist.count("")
        for i in range(n):
            rowlist.remove(u'')
        return rowlist

    def readasdict(self):
        d = {}
        col = self.table.col_values(0)
        nrows = self.table.nrows
        for i in range(nrows):
            val = self.rowlist(i)[1:]
            if len(val) == 1:
                 d[col[i]] = val[0]
            else:
                d[col[i]] = val
        return d

    def readaslitbyrow(self, i, j):
        l = []
        s = self.rowlist(i)
        e = self.rowlist(j)
        for i in range(1, len(s)):
            d = {}
            d[s[0]] = s[i]
            d[e[0]] = e[i]
            l.append(d)
        return l

    def writeCell(self, row, col, content,):
        """
        :param row: 行号
        :param col: 列号
        :param content: 写入内容

        :return:
        """
        #写入单元格

        self.wb = copy(self.data) #复制打开的文件
        ws = self.wb.get_sheet(self.sName)
        ws.write(row, col, content)

    def writeSave(self, s_url=None, is_new="F"):
        """
        :param s_url: 新保存的文件路径
        :param is_new: 是否新建文件
        :return:
        """
        if is_new == "F":
            os.remove(self.cPath)
            self.wb.save(self.cPath)
        else:
            self.wb.save(s_url)
        print("保存成功")

    def copyCol(self, OsheetName, TsheetName, is_origin, col, sPath):
        #复制插入列数据。
        wb = load_workbook(self.cPath)
        ws = wb[OsheetName]
        sh = wb.active
        ar = sh.max_row #行数
        bl = []
        for j in range(0, ar):
            bl.insert(j, ws.cell(j, 2).value)
        if is_origin == True:
            wsO = wb[TsheetName]
            for i in range(0, ar):
                while i < ar:
                    wsO.cell(i + 1, int(col)).value = bl[i]
            wb.save(self.cPath)
            print("保存成功！")
        elif is_origin == False:
            wbT = load_workbook (sPath)
            wsO = wbT[TsheetName]
            for i in range(0, ar):
                wsO.cell(i + 1, int(col)).value = bl[i]
            wb.save(sPath)
            print("保存成功！")
        else:
            print("请输入‘True’或者 ‘False’" )
            return


if __name__=="__main__":
    oper = ExcelUtil("D:\\userinfo.xlsx", "ai")
    info = oper.next()


    # for ins in info:
    #     # oper.writeCell(int(ins["ID"])+1,0,ins["ID"])
    #     # oper.writeCell(int(ins["ID"]) + 1, 1, ins["PHONE"])
    #     # oper.writeCell(int(ins["ID"]) + 1, 2, ins["ACCOUNT_NO"])
    #     # oper.writeCell(int(ins["ID"]) + 1, 3, ins["IDENT"])
    #     # oper.writeCell(int(ins["ID"]) + 1, 4, ins["REGISTER"])
    #     # oper.writeCell(int(ins["ID"]) + 1, 5, ins["REMARK"])
    #     print(ins)

    #oper.writeSave("D:\\userinfo.txt", "T")


