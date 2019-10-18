# coding=utf-8  
"""
@author: mirrorChen
@license: (C) Copyright 2011-2018, mirror personal Limited.
@contact: chenjingxu3@dafycredit.com
@software: JY_Android_AT
@file: u2Main.py
@time: 2019/7/19 11:46
@desc:
"""
import unittest
import time
import random
from project_01.case.mysql_pubtwo import MysqlUtiltwo
from project_01.case.oracle_pub import OracleUtil
from project_01.case.excel_pub import ExcelUtil
from project_01.case.InterfaceClass import APITest
from openpyxl import load_workbook
import uiautomator2 as u2
from project_01.case.API_Para import API_Parameters
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import keywords
import inspect
import os

class Method():
    def __init__(self, driver):
        self.driver = driver

    def findItem(self,el):
        time.sleep(1)
        source = self.driver.page_source
        if el in source:
            return True
        else:
            return False


    def createName(self):
        a1 = ['张', '金', '李', '王', '赵', '陈', '胡', '莫', '卫', '韦', '霍', '贾', '周', '陆']
        a2 = ['玉', '明', '龙', '芳', '军', '玲', '镜', '元', '都', '红', '绑', '问', '有', '那', '及', '公', '的', '龙', '格', '达', '离','家', '个','极','开']
        a3 = ['来', '立', '玲', '', '国', '及', '订', '额', '发', '哥', '试', '还', '吧', '饿', '运', '扣', '流', '考', '岭', '嗯', '玩','解', '密', '点', '来', '放', '哦']
        name = "测试" + random.choice(a1) + random.choice(a2) + random.choice(a3)
        return name

    def insertCell(self, oPath, TsheetName, is_origin, row, col, contens, sPath=None):
        #插入列数据。
        wb = load_workbook(oPath)
        if is_origin == True:
            wsO = wb[TsheetName]
            wsO.cell(int(row), int(col)).value = contens
            wb.save(oPath)
            print("保存成功！")
        elif is_origin == False:
            wbT = load_workbook(sPath)
            wsO = wbT[TsheetName]
            wsO.cell(int(row), int(col)).value = contens
            wbT.save(sPath)
            print("保存成功！")
        else:
            print("请输入‘True’或者 ‘False’" )
            return

    def is_toast_exist(self, driver, toastmessage, timeout=5, poll_frequency=0.5):
        """is toast exist, return True or False
        :Agrs:
         - driver - 传driver
         - toastmessage   - 页面上看到的toast消息文本内容
         - timeout - 最大超时时间，默认30s
         - poll_frequency  - 间隔查询时间，默认0.5s查询一次
        :Usage:
         is_toast_exist(driver, "toast消息的内容")
        """
        try:
            toast_loc = ("xpath", ".//*[contains(@text,'%s')]" % toastmessage)
            WebDriverWait(driver, timeout, poll_frequency).until(
                expected_conditions.presence_of_element_located(toast_loc))
            return True
        except:
            return False

        #element = WebDriverWait(driver,timeout,poll_frequency).until(expected_conditions.presence_of_element_located((By.PARTIAL_LINK_TEXT,message)))
    def toast_judge(self, driver):
        t_msg = driver.toast.get_message(5.0,10.0,"没有获取到toast")
        driver.toast.reset()
        return t_msg


    #获取类方法
    def class_loader(cls):
        _class = getattr(keywords, cls)
        return _class()

    def get_kw_list(self):
        func_list = []
        for m in keywords.__all__:
            _class = self.class_loader(m)
            for name, value in inspect.getmembers(_class):
                if not name.startswith("_"):
                    func_list.append(name)
        return func_list



class createData(unittest.TestCase):

    def setUp(self):
        # 初始化信息
        caseName ="userinfo.xlsx"
        cur_path = os.path.dirname(os.path.realpath(__file__))
        print(cur_path)
        self.f_path = os.path.join(cur_path, caseName)
        ff_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
        self.shotPicPath = os.path.join(ff_path,"shotPics\/")
        print("self.shotPicPath ", self.shotPicPath)
        self.sheetName = "a1"
        self.Excel = ExcelUtil(self.f_path, self.sheetName)
        self.MDB = MysqlUtiltwo()
        self.ODB = OracleUtil()
        self.API = APITest()
        self.API_Para = API_Parameters()

        self.driver = u2.connect("721QEBR925Q45")
        self.Method = Method(self.driver)
        self.driver.wait_timeout = 3.0

    def test_createData(self):
        #造数据
        i = 0
        info = self.Excel.next()
        j = len(info) - 1
        registered = []  # 记录已注册用户
        while j >= 0:
            if info[j]["REGISTER"] == "注册成功":
                registered.append(info.pop(j))  # 已注册的则剔除,并记录到一个列表中
            j = j - 1
        dev_info = self.driver.device_info
        width = dev_info["display"]["width"]  # 获取屏幕的宽度
        height = dev_info["display"]['height']  # 获取屏幕的高度


        while i < len(info):
            #try:
            self.driver.app_start("com.giveu.shoppingmall")
            self.ID = info[i]["ID"]
            self.PHONE = info[i]["PHONE"]
            self.ACCOUNT_NO = info[i]["ACCOUNT_NO"]
            self.IDENT = info[i]["IDENT"]
            self.REGISTER = info[i]["REGISTER"]
            self.REMARK = info[i]["REMARK"]

            uName = self.Method.createName()
            print("提单人姓名：", uName)
            print("提单手机号：", self.PHONE)
            print("提单身份证：", self.IDENT)
            print("提单银行卡：", self.ACCOUNT_NO)

            # 造数据：到信用授权通过.
            time.sleep(2)
            try:
                #self.driver(text="取消").click()
                self.driver(textContained="允许").click()
                self.driver(textContained="允许").click()
                print("授权成功")
            except:
                print("无需授权")
            try:
                self.driver(resourceId="com.giveu.shoppingmall:id/tv_skip").click()
            except:
                self.driver.swipe(width / 2, height / 2, 0, height / 2)
                time.sleep(0.1)
                self.driver.swipe(width / 2, height / 2, 0, height / 2)
                time.sleep(0.1)
                self.driver.swipe(width / 2, height / 2, 0, height / 2)
                time.sleep(0.1)
                self.driver.swipe(width / 2, height / 2, 0, height / 2)
                time.sleep(0.2)
                self.driver(text="立即体验").click()
            try:
                self.driver(text="切换账号").click()
                self.driver(resourceId="com.giveu.shoppingmall:id/iv_back").click()
                print("成功跳转登录页")
                time.sleep(2)
            except:
                print("登录页异常")
            # 注册
            # 注册 --跳转注册页面
            print("开始注册")
            self.driver(resourceId="com.giveu.shoppingmall:id/bottom_icon", instance=3).click() # 点击我的
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_login").click()
            #self.driver(resourceId="com.giveu.shoppingmall:id/xy_oauth_back").click()  #一键登录
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_register").click()
            time.sleep(0.3)
            # 注册--注册账号页面
            print("注册账号")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_content", instance=0).send_keys(self.PHONE)
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_send_code").click()
            time.sleep(2)  # 等两秒再去填写验证码
            self.driver(resourceId="com.giveu.shoppingmall:id/et_content", instance=1).send_keys("123456")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_next").click()
            time.sleep(0.5)
            # 注册--设置密码页面
            print("设置密码")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_content", instance=0).send_keys("abc12345")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_content", instance=1).send_keys("abc12345")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_complete").click()
            self.driver.toast.reset()
            r_msg = self.driver.toast.get_message(5, 10, "没有获取到注册成功信息")
            if r_msg == "注册成功":
                print(r_msg)
                # EXCEL中标记注册成功
                self.Method.insertCell(self.f_path, self.sheetName, True, int(self.ID) + 1, 5, "注册成功")
            else:
                self.assertEqual(r_msg, "注册成功", "注册失败！")
                self.driver.app_clear()
            i = i + 1
            #time.sleep(3)
            # 登录  --注册之后不需要填写账号
            print("登录账号")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_content", instance=1).send_keys("abc12345")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_login").click()
            print("登录成功！")
            time.sleep(1.5)
            self.driver(text="取消").click()
            sql = "select access_token from t_user_token tu where user_name='" + str(self.PHONE) + "' order by create_time desc limit 1"
            self.tk = self.MDB.mysql_getstring(sql)

            # 调用接口，插入活体图片数据，跳过活体
            sql = "select id from t_user where login_name=" + str(self.PHONE)
            userid = self.MDB.mysql_getstring(sql)
            postparams = self.API_Para.skiplivingbody(userid)
            getparams = ''
            user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
            tokens = 'Authorization: Bearer ' + str(self.tk)
            # 添加头信息
            headers = {'User-Agent': user_agent,
                       'Content-Type': 'application/json',
                       'token': tokens
                       }
            retInfo = self.API.apicall("POST",
                                       "http://testdafyshop.dafycredit.cn/v1/cashloan/applyInfoRecord/substepSave",
                                       getparams,
                                       postparams,
                                       headers
                                       )
            print(retInfo)

            # 取现部分----------------------------------------------------------------------------------------------------
            print("进入取现流程")
            self.driver(resourceId="com.giveu.shoppingmall:id/bottom_icon", instance=0).click()
            self.driver.xpath('//*[@resource-id="com.giveu.shoppingmall:id/bannerViewPager"]/android.widget.ImageView[1]').click()
            time.sleep(0.5)
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_apply", instance=0).click()  # 点击立即申请
            self.driver(text="活体检测").click()
            # 取现部分 --绑定身份证信息
            # 正面
            print("身份证正面")
            self.driver(resourceId="com.giveu.shoppingmall:id/iv_pocket_money_add", instance=0).click()
            self.driver(resourceId="com.meizu.media.camera:id/shutter_btn").click()
            self.driver(resourceId="com.meizu.media.camera:id/btn_done").click()
            # 反面
            time.sleep(3)
            print("身份证反面")
            self.driver.swipe(460, 950, 460, 200)
            time.sleep(0.5)
            self.driver(resourceId="com.giveu.shoppingmall:id/iv_pocket_money_add").click()
            time.sleep(0.5)
            self.driver(resourceId="com.meizu.media.camera:id/shutter_btn").click()  # 魅族note6
            self.driver(resourceId="com.meizu.media.camera:id/btn_done").click()  # 魅族note6

            # 手持
            time.sleep(3)
            print("手持身份证")
            self.driver.swipe(460, 950, 460, 200)
            time.sleep(0.5)
            self.driver(resourceId="com.giveu.shoppingmall:id/iv_pocket_money_add").click()
            time.sleep(0.5)
            self.driver(resourceId="com.meizu.media.camera:id/shutter_btn").click()  # 魅族note6
            self.driver(resourceId="com.meizu.media.camera:id/btn_done").click()  # 魅族note6
            # 填写身份证信息
            time.sleep(3)
            print("身份证信息")
            self.driver.swipe(460, 950, 460, 150)
            time.sleep(0.5)
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=0).clear_text()  # 姓名
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=0).send_keys(uName)
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=1).clear_text()
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=1).send_keys(
                self.IDENT)  # 身份证号码
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=2).clear_text()
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=2).send_keys(
                "福中一路地铁大厦17楼")  # 户籍地址
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=3).clear_text()
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=3).send_keys(
                "福田区公安局")  # 发证机关
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=0).click()
            time.sleep(0.3)
            self.driver(text="京族").click()  # 民族
            time.sleep(0.3)
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=1).click()  # 身份证地址
            time.sleep(0.3)
            self.driver(text="福建").click()
            self.driver(text="莆田市").click()
            self.driver(text="仙游县").click()
            time.sleep(0.3)
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=2).click()  # 有效日期起

            self.driver.swipe(180, 1384, 180, 1700)
            time.sleep(0.5)
            self.driver(text="确定").click()
            self.driver.swipe(460, 950, 460, 250)
            time.sleep(0.5)
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=3).click()  # 有效日期止
            time.sleep(0.3)
            self.driver.swipe(180, 1384, 180, 1000)
            time.sleep(0.3)
            self.driver(text="确定").click()
            time.sleep(2)  # 这需要2秒，不然会有失败

            self.driver(resourceId="com.giveu.shoppingmall:id/tv_next").click()


            # 基本信息
            # 个人
            print("个人信息")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=0).click()# 居住地址
            time.sleep(0.3)
            self.driver(text="福建").click()
            self.driver(text="莆田市").click()
            self.driver(text="仙游县").click()
            time.sleep(0.3)
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=0).send_keys("福中一路地铁大厦17楼")  # 详细地址
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=1).click()# 学历
            time.sleep(0.3)
            self.driver(text="小学").click()
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=1).send_keys("123456@qq.com")  # 邮箱
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=2).click() # 婚姻状态
            time.sleep(0.3)
            self.driver(text="未婚").click()
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=3).click()# 月收入
            self.driver(text="3000-4000").click()
            #家庭月收入
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=4).click()  # 月收入
            self.driver(text="12000-16000").click()
            self.driver.swipe(460, 1080, 460, 150)
            time.sleep(0.5)
            # 公司
            print("公司信息")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=0).click()  # 月收入
            #self.driver(text="河北").click()
            #self.driver(text="衡水市").click()
            self.driver(text="仙游县", instance=1).click()

            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=0).send_keys(
                "深圳福田哈哈村某栋某层")  # 公司地址
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=1).send_keys(
                "深圳即融科技有限公司")  # 公司名称
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=1).click() #职位
            self.driver(text="司机").click()
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=2).send_keys("0755-86511248")# 公司电话
            self.driver.swipe(460, 1080, 460, 150)
            time.sleep(0.5)
            # 家庭联系人
            print("家庭联系人信息")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=1).send_keys("熊大")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=2).send_keys("13410342889")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=0).click()  # 与本人关系
            self.driver(text="兄弟").click()
            # 其它联系人
            print("其它联系人信息")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=3).send_keys("熊二")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=4).send_keys("13410342879")
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_text_right", instance=1).click()# 与本人关系
            self.driver(text="同学").click()
            # 提交下一步
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_next").click()
            time.sleep(1)

            # 绑定银行卡
            print("绑定银行卡")
            # self.driver.activate_ime_engine("com.baidu.input_huawei/.ImeService")
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=1).send_keys(self.ACCOUNT_NO)
            time.sleep(1.5)
            self.driver(resourceId="com.giveu.shoppingmall:id/et_text_right", instance=2).send_keys(self.PHONE)  # 预留手机号
            time.sleep(0.5)
            self.driver(text="获取验证码").click()
            time.sleep(2)

            self.driver(resourceId="com.giveu.shoppingmall:id/et_send_code").send_keys("123456")
            self.driver.click(964, 1221)
            # 提交下一步
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_next").click()

            time.sleep(1.5)
            print("绑定银行卡成功")
            # 信用提升
            print("信用提升")
            self.driver(resourceId="com.giveu.shoppingmall:id/bt_credit", instance=0).click()
            time.sleep(3)
            sql = "update dafy_sales.EXTERNAL_FS_TASK set task_id='fsd4354332jk3123d',is_get_report=1 where name='" + uName + "'"
            self.ODB.oracle_sql(sql)
            time.sleep(1)
            self.driver(resourceId="com.giveu.shoppingmall:id/top_left_image").click()
            #self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_refresh").click()  # 刷新
            time.sleep(1)
            self.driver(resourceId="com.giveu.shoppingmall:id/tv_next").click()
            #这里需要获取toast，因为不成功会弹出toast。
            t_msg_1="请完成所有授权后进行提交"
            if self.Method.is_toast_exist(self.driver,t_msg_1):
                self.driver(resourceId="com.giveu.shoppingmall:id/bt_credit", instance=0).click()
                time.sleep(3)
                sql = "update dafy_sales.EXTERNAL_FS_TASK set task_id='fsd4354332jk3123d',is_get_report=1 where name='" + uName + "'"
                self.ODB.oracle_sql(sql)
                time.sleep(1)
                self.driver(resourceId="com.giveu.shoppingmall:id/top_left_image").click()
                # self.driver.find_element_by_id("com.giveu.shoppingmall:id/tv_refresh").click()  # 刷新
                time.sleep(1)
                self.driver(resourceId="com.giveu.shoppingmall:id/tv_next").click()

            # 记录授权成功
            t_msg = "分布保存失败，调用合同中心生成合同失败"
            if self.Method.is_toast_exist(self.driver,t_msg):
                self.Method.insertCell(self.f_path, self.sheetName, True, int(self.ID) + 1, 7, "授权失败")
            else:
                self.Method.insertCell(self.f_path, self.sheetName, True, int(self.ID) + 1, 7, "授权成功")
            print("第 " + str(i) + " 个账号数据已创建成功")
            img_success=self.shotPicPath + str(self.PHONE)+".png"
            self.driver.screenshot(img_success)
            self.driver.app_clear("com.giveu.shoppingmall")


            m_sql="select id_person from t_user where login_name=" + str(self.PHONE)+" limit 1"
            idPerson = self.MDB.mysql_getstring(m_sql)

            o_sql1 = "select id from dafy_sales.cs_credit a where id_person=" + str(idPerson) + " order by app_date desc"
            idCredit = self.ODB.oracle_getstring(o_sql1)

            # 手动决策（赋予额度，让流程跑下去）
            headers = {'User-Agent': user_agent,
                       'Content-Type': 'application/json',
                       }
            postparams = self.API_Para.handDecisionCredit(idCredit)
            d_hdc = self.API.apicall("POST",
                             "http://10.10.11.132:9912/api/Decision/HandDecisionCredit",
                             getparams,
                             postparams,
                             headers)
            print(d_hdc)
            if d_hdc["status"] == True:
                print("决策成功")
            else:
                self.assertEqual(d_hdc["$id"], "2", u"手动决策不成功！")
            # 如果用户没有额度会进入贷超页面，所以进入流程前需判断用户是否有额度，没有则需添加
            o_sql2="select credit_limit from dafy_sales.giveu_online_limit where id_person=" + str(idPerson)
            creditLimit = self.ODB.oracle_getstring(o_sql2)  #返回一个列表
            cl_info = ''
            if creditLimit != "":
                print("得到的额度是：", creditLimit)
            else:
                print("没有获取到额度，执行手动添加额度")
                u_sql = "update dafy_sales.giveu_online_limit set credit_limit=3000 where id_person=" + str(idPerson)
                self.ODB.oracle_sql(u_sql)
                cl_info = self.ODB.oracle_getstring(o_sql2)
                print("手动设置的额度是：",cl_info)

            #借钱
            #self.driver(resourceId="com.giveu.shoppingmall:id/examine_status_btn").click()
            #self.driver(resourceId="com.giveu.shoppingmall:id/borrow_money").clear_text()   #借款金额
            #self.driver(xpath="//*[@resource-id='com.giveu.shoppingmall:id/gv_staging_type_borrow']/android.widget.LinearLayout[1]")  #6期
            #self.driver(xpath='//*[@text="还款详情"]').click() #借款详情
            #self.driver(resourceId="com.giveu.shoppingmall:id/submit_borrow").click()  #提交
            #self.driver(text="前往查看").click()
            #except Exception as e:
            img_fail = self.shotPicPath + str(self.PHONE) + ".png"
            self.driver.screenshot(img_fail)
                #print(e)
            self.driver.app_stop("com.giveu.shoppingmall")

    def tearDown(self):
        self.driver.app_stop("com.giveu.shoppingmall")



if __name__ == "__main__":
    unittest.main()



